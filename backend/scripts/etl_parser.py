import sys
sys.path.append(r"C:\Users\jhenr\AppData\Roaming\Python\Python314\site-packages")
sys.path.append(r"C:\Users\jhenr\AppData\Local\Programs\Python\Python314\Lib\site-packages")

import json
import csv
import os
import re

def clean_decimal(val):
    if val is None:
        return 0.0
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    # Eliminar $, comas de miles, y espacios
    val_str = val_str.replace('$', '').replace(' ', '')
    # Si hay comas y puntos, e.g. 1,662.57
    if ',' in val_str and '.' in val_str:
        val_str = val_str.replace(',', '')
    elif ',' in val_str:
        parts = val_str.split(',')
        if len(parts) == 2 and len(parts[1]) == 2: # probable decimal
            val_str = val_str.replace(',', '.')
        else:
            val_str = val_str.replace(',', '')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def is_numeric(val):
    try:
        float(str(val).replace('$', '').replace(',', '').replace(' ', '').strip())
        return True
    except ValueError:
        return False

# Lista de palabras clave para secciones de vehículos
SECTION_KEYWORDS = [
    'MOTOCICLETAS', 'AUTOMOVILES', 'JEEP', 'MINIVAN', 'VANETTES', 'CAMIONETAS', 'FURGONETAS',
    'AUTUBUS', 'AUTOBUS', 'CAMIONES', 'CAMION', 'MAQUINAS', 'PLACA', 'REMOLQUES', 'VEHICULOS'
]

def analyze_section_header(text):
    text_upper = text.upper().strip()
    
    # Determinar Uso
    use = 'PRIVADO'
    if 'PUBLICO' in text_upper:
        use = 'PUBLICO'
    elif 'RENT' in text_upper or 'TAXI' in text_upper or 'ALQUILER' in text_upper:
        use = 'RENT CAR'
        
    # Determinar Tipo de Vehículo
    vehicle_type = None
    if 'MOTOCICLETAS' in text_upper:
        vehicle_type = 'MOTOCICLETAS'
    elif 'AUTOMOVILES' in text_upper or 'AUTOMOVIL' in text_upper:
        vehicle_type = 'AUTOMOVILES'
    elif 'JEEP' in text_upper:
        vehicle_type = 'JEEP'
    elif 'MINIVAN' in text_upper or 'VANETTES' in text_upper:
        vehicle_type = 'MINIVAN / VANETTES'
    elif 'CAMIONETAS' in text_upper and 'FURGONETAS' in text_upper:
        vehicle_type = 'CAMIONETAS_Y_FURGONETAS'
    elif 'CAMIONETAS' in text_upper:
        vehicle_type = 'CAMIONETAS'
    elif 'FURGONETAS' in text_upper:
        vehicle_type = 'FURGONETAS'
    elif 'AUTUBUS' in text_upper or 'AUTOBUS' in text_upper:
        vehicle_type = 'AUTOBUS'
    elif 'CAMA O CAJA' in text_upper or ('CAMION' in text_upper and 'VOLTEO' not in text_upper and 'CABEZOTE' not in text_upper and 'PATANA' not in text_upper and 'TRACTOR' not in text_upper):
        vehicle_type = 'CAMION (CAMA O CAJA)'
    elif 'VOLTEO' in text_upper:
        vehicle_type = 'CAMIONES VOLTEO'
    elif 'CABEZOTE' in text_upper or 'PATANA' in text_upper or 'TRACTOR' in text_upper:
        vehicle_type = 'CAMIONES CABEZOTE, TRACTOR O PATANA'
    elif 'MAQUINAS' in text_upper or 'AUTOPROPULSION' in text_upper:
        vehicle_type = 'MAQUINAS PESADAS CON AUTOPROPULSION'
    elif 'PLACA' in text_upper or 'DEMOSTRACION' in text_upper:
        vehicle_type = 'PLACA DE DEMOSTRACION'
    elif 'REMOLQUES' in text_upper:
        vehicle_type = 'REMOLQUES'
        
    return vehicle_type, use

def analyze_file(file_path):
    if not os.path.exists(file_path):
        return {'exito': False, 'mensaje': f'El archivo no existe: {file_path}'}
        
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.csv':
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                sample = f.read(2048)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t'])
                reader = csv.reader(f, dialect)
                rows = list(reader)
                
                if not rows:
                    return {'exito': False, 'mensaje': 'El archivo CSV está vacío.'}
                    
                headers = [h.strip() for h in rows[0]]
                preview = []
                for r in rows[1:4]:
                    preview.append(r)
                    
                return {
                    'exito': True,
                    'type': 'csv',
                    'sheets': [],
                    'columns': headers,
                    'preview': preview
                }
        except Exception as e:
            return {'exito': False, 'mensaje': f'Error al leer CSV: {str(e)}'}
            
    elif ext in ['.xlsx', '.xls']:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            if not sheets:
                return {'exito': False, 'mensaje': 'El archivo Excel no contiene hojas.'}
                
            # Buscar una hoja con datos reales (generalmente la primera o alguna que tenga cabeceras)
            sheet = wb[sheets[0]]
            
            # Buscar una fila de cabecera con nombres reales (ej. fila 12 en Multiseguros)
            all_rows = []
            for r in sheet.iter_rows(values_only=True):
                all_rows.append(list(r))
                if len(all_rows) >= 50:
                    break
                    
            if not all_rows:
                return {'exito': False, 'mensaje': 'La primera hoja del Excel está vacía.'}
                
            # Heurística para buscar la fila de cabecera:
            # Es la primera fila que contiene palabras de tarifas como 'MULTISEGUROS' o 'INTENSION' o tiene varios textos no vacíos
            header_row_idx = 0
            for i, r in enumerate(all_rows):
                r_strs = [str(cell).upper().strip() for cell in r if cell is not None]
                if any(x in y for y in r_strs for x in ['MULTISEGUROS', 'INTENSION', 'TARIFAS', 'PRECIO', 'PRIMA']):
                    header_row_idx = i
                    break
            
            headers = []
            for h in all_rows[header_row_idx]:
                if h is not None:
                    headers.append(str(h).strip())
                else:
                    headers.append("")
                    
            # Si no encontramos cabecera por texto, tomamos la fila 0
            if not any(headers):
                headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
                header_row_idx = 0
                
            preview = []
            data_start_idx = header_row_idx + 1
            for r in all_rows[data_start_idx:data_start_idx+10]:
                preview.append([str(cell) if cell is not None else "" for cell in r])
                
            return {
                'exito': True,
                'type': 'xlsx',
                'sheets': sheets,
                'columns': headers,
                'preview': preview,
                'header_row_index': header_row_idx
            }
        except Exception as e:
            return {'exito': False, 'mensaje': f'Error al leer Excel: {str(e)}'}
    else:
        return {'exito': False, 'mensaje': f'Extensión de archivo no soportada: {ext}'}

def parse_file(file_path, mapping_json, sheet_name=None):
    if not os.path.exists(file_path):
        return {'exito': False, 'mensaje': f'El archivo no existe: {file_path}'}
        
    try:
        import base64
        # Intentar decodificar como base64 por si viene de PHP en Windows
        try:
            decoded = base64.b64decode(mapping_json).decode('utf-8')
            if decoded.strip().startswith('{') and decoded.strip().endswith('}'):
                mapping_json = decoded
        except Exception:
            pass
            
        mapping = json.loads(mapping_json)
    except Exception as e:
        return {'exito': False, 'mensaje': f'Error al parsear el JSON de mapeo: {str(e)}'}
        
    ext = os.path.splitext(file_path)[1].lower()
    raw_rows = []
    
    if ext == '.csv':
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                sample = f.read(2048)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t'])
                reader = csv.reader(f, dialect)
                raw_rows = list(reader)
        except Exception as e:
            return {'exito': False, 'mensaje': f'Error al leer CSV: {str(e)}'}
            
    elif ext in ['.xlsx', '.xls']:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if not sheet_name or sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            
            for r in sheet.iter_rows(values_only=True):
                raw_rows.append(list(r))
        except Exception as e:
            return {'exito': False, 'mensaje': f'Error al leer Excel: {str(e)}'}
            
    # Determinar modo de parseo: bloques (jerárquico) o plano
    mode = mapping.get('mode', 'flat')
    parsed_records = []
    errors = []
    
    if mode == 'blocks':
        # --- PARSEO EN MODO BLOQUES (JERÁRQUICO) ---
        current_type = None
        current_use = 'PRIVADO'
        
        # En modo bloques, el mapeo define en qué columnas buscar la capacidad y la tarifa base
        cap_col_name = mapping.get('capacidad', {}).get('value', '')
        tarifa_col_name = mapping.get('tarifa_base', {}).get('value', '')
        
        # Encontrar los índices de estas columnas a partir de la fila de cabecera de referencia
        # (usualmente la fila con cabecera)
        header_row_idx = 0
        for i, r in enumerate(raw_rows):
            r_strs = [str(cell).upper().strip() for cell in r if cell is not None]
            if any(x in y for y in r_strs for x in [tarifa_col_name.upper(), 'MULTISEGUROS', 'INTENSION']):
                header_row_idx = i
                break
                
        headers = [str(h).strip() if h is not None else "" for h in raw_rows[header_row_idx]]
        
        cap_idx = headers.index(cap_col_name) if cap_col_name in headers else 0
        tarifa_idx = headers.index(tarifa_col_name) if tarifa_col_name in headers else 3
        
        for idx, row in enumerate(raw_rows):
            row_num = idx + 1
            if row_num <= header_row_idx:
                continue # Omitir filas antes de la cabecera
                
            # Limpiar fila
            row_cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(row_cells):
                continue # Omitir fila vacía
                
            col_0 = row_cells[0]
            
            # Detectar si es cabecera de sección/bloque
            # Regla: Si Column 0 contiene alguna keyword de sección y la columna de tarifa está vacía o no es numérica
            val_tarifa_raw = row_cells[tarifa_idx] if tarifa_idx < len(row_cells) else ""
            
            is_section = False
            if col_0 and not is_numeric(val_tarifa_raw):
                col_0_upper = col_0.upper()
                if any(k in col_0_upper for k in SECTION_KEYWORDS):
                    is_section = True
                    
            if is_section:
                # Extraer tipo de vehículo y uso
                v_type, use = analyze_section_header(col_0)
                if v_type:
                    current_type = v_type
                    current_use = use
                continue
                
            # Si es fila de datos y tenemos un tipo de vehículo detectado
            if current_type and col_0 and is_numeric(val_tarifa_raw):
                capacidad = col_0
                tarifa_val = clean_decimal(val_tarifa_raw)
                
                if tarifa_val <= 0:
                    continue
                    
                # Determinar Cobertura por defecto según tipo de vehículo y capacidad
                cobertura = 'LIVIANO BASICO'
                if current_type == 'MOTOCICLETAS':
                    if any(x in capacidad for x in ['250', '350', 'Eléctrica', 'Patineta']):
                        cobertura = 'MOTOCICLETA BASICO'
                    else:
                        cobertura = 'LIVIANO BASICO'
                elif current_type in ['AUTOBUS', 'CAMION (CAMA O CAJA)', 'CAMIONES VOLTEO', 'CAMIONES CABEZOTE, TRACTOR O PATANA', 'MAQUINAS PESADAS CON AUTOPROPULSION', 'PLACA DE DEMOSTRACION', 'REMOLQUES']:
                    cobertura = 'PESADO PLUS'
                    
                # Si el tipo es la combinación "CAMIONETAS_Y_FURGONETAS", creamos un registro para cada una
                if current_type == 'CAMIONETAS_Y_FURGONETAS':
                    # Camioneta
                    parsed_records.append({
                        'tipo_vehiculo': 'CAMIONETAS',
                        'capacidad': capacidad,
                        'uso': current_use,
                        'cobertura': cobertura,
                        'tarifa_base': tarifa_val
                    })
                    # Furgoneta (solo si la capacidad no contiene "5 Pasajeros" o "Platanera" que son exclusivas)
                    if not any(x in capacidad for x in ['5 Pasajeros', 'Platanera']):
                        # En furgonetas, la capacidad de "Hasta 2 Tonelada y/o 5 Pasajeros" se llama "Hasta 2 Tonelada"
                        cap_furgoneta = capacidad
                        if 'Hasta 2 Tonelada' in capacidad:
                            cap_furgoneta = 'Hasta 2 Tonelada'
                        parsed_records.append({
                            'tipo_vehiculo': 'FURGONETAS',
                            'capacidad': cap_furgoneta,
                            'uso': current_use,
                            'cobertura': cobertura,
                            'tarifa_base': tarifa_val
                        })
                else:
                    parsed_records.append({
                        'tipo_vehiculo': current_type,
                        'capacidad': capacidad,
                        'uso': current_use,
                        'cobertura': cobertura,
                        'tarifa_base': tarifa_val
                    })
                    
    else:
        # --- PARSEO EN MODO TABLA PLANA ---
        if not raw_rows:
            return {'exito': False, 'mensaje': 'Archivo vacío.'}
            
        headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
        
        for idx, row in enumerate(raw_rows[1:]):
            row_num = idx + 2
            row_cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(row_cells):
                continue
                
            record = {}
            row_valid = True
            
            for target_col, map_cfg in mapping.items():
                if target_col == 'mode':
                    continue
                if map_cfg.get('type') == 'fixed':
                    record[target_col] = map_cfg.get('value')
                elif map_cfg.get('type') == 'column':
                    col_name = map_cfg.get('value')
                    if col_name in headers:
                        col_idx = headers.index(col_name)
                        if col_idx < len(row_cells):
                            record[target_col] = row_cells[col_idx]
                        else:
                            record[target_col] = ""
                    else:
                        record[target_col] = ""
                        if target_col in ['tipo_vehiculo', 'tarifa_base']:
                            errors.append(f"Fila {row_num}: Columna mapeada '{col_name}' no encontrada.")
                            row_valid = False
                else:
                    record[target_col] = ""
                    
            if not row_valid:
                continue
                
            # Limpiar e normalizar
            record['tipo_vehiculo'] = str(record.get('tipo_vehiculo', '')).strip().upper()
            record['capacidad'] = str(record.get('capacidad', '')).strip()
            
            # Normalizar Uso
            uso_raw = str(record.get('uso', '')).strip().upper()
            if any(term in uso_raw for term in ['PRIV', 'PRIVADO', 'PV']):
                record['uso'] = 'PRIVADO'
            elif any(term in uso_raw for term in ['PUB', 'PUBLICO', 'PB']):
                record['uso'] = 'PUBLICO'
            elif any(term in uso_raw for term in ['RENT', 'ALQUILER', 'RC']):
                record['uso'] = 'RENT CAR'
            else:
                record['uso'] = uso_raw if uso_raw else 'PRIVADO'
                
            record['cobertura'] = str(record.get('cobertura', '')).strip().upper()
            
            # Autorellenar cobertura si está vacía
            if not record['cobertura']:
                tipo = record['tipo_vehiculo']
                cap = record['capacidad']
                if tipo == 'MOTOCICLETAS':
                    if any(x in cap for x in ['250', '350', 'Eléctrica', 'Patineta']):
                        record['cobertura'] = 'MOTOCICLETA BASICO'
                    else:
                        record['cobertura'] = 'LIVIANO BASICO'
                elif tipo in ['AUTOMOVILES', 'JEEP', 'MINIVAN / VANETTES', 'CAMIONETAS', 'FURGONETAS']:
                    record['cobertura'] = 'LIVIANO BASICO'
                else:
                    record['cobertura'] = 'PESADO PLUS'
                    
            record['tarifa_base'] = clean_decimal(record.get('tarifa_base', 0.0))
            
            if not record['tipo_vehiculo']:
                errors.append(f"Fila {row_num}: El tipo de vehículo está vacío.")
                continue
            if record['tarifa_base'] <= 0:
                errors.append(f"Fila {row_num}: Tarifa base inválida o cero.")
                continue
                
            parsed_records.append(record)
            
    return {
        'exito': True,
        'datos': parsed_records,
        'errores': errors,
        'total_procesados': len(raw_rows),
        'total_validos': len(parsed_records)
    }

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({'exito': False, 'mensaje': 'Uso: etl_parser.py <cmd> <file_path> [args]'}))
        sys.exit(1)
        
    cmd = sys.argv[1]
    file_path = sys.argv[2]
    
    if cmd == 'analyze':
        res = analyze_file(file_path)
    elif cmd == 'parse':
        if len(sys.argv) < 4:
            res = {'exito': False, 'mensaje': 'Falta el JSON de mapeo para procesar el archivo.'}
        else:
            mapping_json = sys.argv[3]
            sheet_name = sys.argv[4] if len(sys.argv) > 4 else None
            res = parse_file(file_path, mapping_json, sheet_name)
    else:
        res = {'exito': False, 'mensaje': f'Comando desconocido: {cmd}'}
        
    print(json.dumps(res, indent=2, default=str))
